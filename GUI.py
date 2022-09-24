"""
項目分工:
    LIN:
    添加UI:
        無須一定要啟動Camera也能訓練模型
        添加訊息框(訓練完成, 歡迎xx同學 等等...)
    訓練模型模塊:
                原本使用opencv掃臉模型改使用mediapipe識別截圖
                添加目前拍到第幾張 and {學號}
    識別模型模塊:
                將原本的opencv改成用mediapipe截臉部
    解決剛開始沒掃到臉崩潰的問題
    增加識別容錯度
    增加標記目前該同學資料(照片)數
    -----
    BOTI:
    添加點名系統模塊:
        添加資料庫
        將名字寫入資料庫
        自動辨別月份並只紀錄平日
        掃臉成功識別後紀錄出席
        
如有BUG麻煩回報
"""

import sys
import tkinter as tk
import tkinter.ttk as ttk
from tkinter.constants import *
from tkinter import messagebox
from ttkthemes import ThemedStyle
import os.path
import cv2
import os
import glob
import numpy as np
import mediapipe as mp
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from datetime import date
from datetime import datetime
import datetime
import os
import calendar
from time import time,localtime
import tkinter as tk
from tkinter import filedialog

import shutil


_script = sys.argv[0]
_location = os.path.dirname(_script)
_bgcolor = '#d9d9d9'  # X11 color: 'gray85'
_fgcolor = '#000000'  # X11 color: 'black'
_compcolor = 'gray40' # X11 color: #666666
_ana1color = '#c3c3c3' # Closest X11 color: 'gray76'
_ana2color = 'beige' # X11 color: #f5f5dc
_tabfg1 = 'black' 
_tabfg2 = 'black' 
_tabbg1 = 'grey75' 
_tabbg2 = 'grey89' 
_bgmode = 'light' 
today = date.today()
d1 = today.strftime("%d")

def student_ID():
    student_ID = tk.Tk()
    student_ID.title('輸入學號')
    student_ID.geometry("300x80+750+350")
    def getTextInput():
        result=textExample.get(1.0, tk.END+"-1c")
        f = open('student_ID.txt', 'w')
        f.write(result)
        f.close()
        student_ID.quit()

    textExample=tk.Text(student_ID, height=1)
    textExample.pack()
    btnRead=tk.Button(student_ID, height=1, width=10, text="Read", 
                        command=getTextInput)
    btnRead.pack()
    student_ID.after(15000,student_ID.destroy)
    student_ID.mainloop()
def excel():
    i=0
    day=0
    ####################################當月月數
    tonow = datetime.datetime.now()
    yyyy=tonow.year
    mm=tonow.month
    dd=tonow.day
    today = date.today()
    d1 = today.strftime("%d")
    days=calendar.monthrange(yyyy,mm)[1]
    ####################################新增excel
    pp="test.xlsx"
    if not os.path.isfile(pp):     
        wb = openpyxl.Workbook()# 利用 Workbook 建立一個新的工作簿
        wb.create_sheet("點名表", 0) # 新增工作表並指定放置位置
        sheet = wb.worksheets[0]# 取得第一個工作表
        sheet[chr(65)+'1'].value = '學號'    #創建基本欄位
        while day<days:
            day=day+1
            temp = pd.Timestamp(str(yyyy)+'-'+str(mm)+'-'+str(day))
            i=i+1
            print('i=',i)
            
            if temp.day_name()=='Saturday':#跳過假日
                day+=2
                if day>days:
                    break
            sheet[chr(65+int(i))+'1'].value =  day
        wb.save(pp)# 儲存檔案
        print('工做表:',wb.sheetnames)      
def Train():
    global name
    total = 5
    mp_face_detection = mp.solutions.face_detection   # 建立偵測方法
    mp_drawing = mp.solutions.drawing_utils           # 建立繪圖方法
    if not os.path.exists("library"):                    # 如果不存在library資料夾
        messagebox.showinfo('提示','文件夾中不存在library現在將自動創建\n請重新點[加入人臉模型]')
        os.mkdir("library")                              # 就建立library
    else:
        inputyml = tk.messagebox.askquestion('錄製人臉','是否啟動攝像頭 (y/n)')
        if inputyml == 'yes':
            student_ID()
            path = 'student_ID.txt'
            f = open(path, 'r')
            name = f.read()
            f.close()
            fileTest = r"student_ID.txt"
            os.remove(fileTest)
            if os.path.exists("library\\" + name):
                messagebox.showinfo('提示','此名字的人臉資料已經存在')
                print("此名字的人臉資料已經存在")
            else:
                os.mkdir("library\\" + name)
                cap = cv2.VideoCapture(0)                       # 開啟攝影機
                num = 1                                         # 影像編號
                with mp_face_detection.FaceDetection(             # 開始偵測人臉
                model_selection=1, min_detection_confidence=0.5) as face_detection:

                    if not cap.isOpened():
                        print("Cannot open camera")
                        exit()
                    while num <= total:
                        ret, img = cap.read()
                        if not ret:
                            print("Cannot receive frame")
                            break
                        else:
                            img2 = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)   # 將 BGR 顏色轉換成 RGB
                            results = face_detection.process(img2)        # 偵測人臉

                            if results.detections:
                                for detection in results.detections:
                                    bboxC = detection.location_data.relative_bounding_box
                                    ih, iw, ic = img.shape
                                    bbox = (int(bboxC.xmin * iw), int(bboxC.ymin * ih), 
                                            int(bboxC.width * iw), int(bboxC.height * ih))
                                    cv2.rectangle(img, bbox, (255,255,255), 1)  # 標記人臉
                                    cv2.putText(img, f"Name: {name} Now{num}", (bbox[0], bbox[1]-20),cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255,0), 1, cv2.LINE_AA)
                                    cv2.imshow('oxxostudio', img)
                                    x = int(bboxC.xmin * iw)
                                    y = int(bboxC.ymin * ih)
                                    w = int(bboxC.width * iw)
                                    h = int(bboxC.height * ih) 
                                    imageCrop = img[y:y+h,x:x+w]                     # 裁切
                                    imageResize = cv2.resize(imageCrop,(160,160))     # 重製大小
                                    faceName = "library\\" + name + "\\" + name +\
                                                str(num) + ".jpg"
                                    cv2.imwrite(faceName, imageResize)      # 儲存人臉影像   
                                    key = cv2.waitKey(200)
                                    print(f"拍攝第 {num} 次人臉成功")
                                    num += 1
                            else:
                                print("請露出人臉或者摘掉遮蔽物!")
                cap.release()
                cv2.destroyAllWindows()
        else:
            print("將以存在的資料夾進行訓練")


    # 讀取人臉樣本和放入faces_db, 同時建立標籤與人名串列
    nameList = []                                       # 學生學號
    faces_db = []                                       # 儲存所有人臉
    labels = []                                         # 建立人臉標籤
    index = 0                                           # 學生編號索引
    dirs = os.listdir('library')                         # 取得所有資料夾及檔案
    for d in dirs:                                      # d是所有學生人臉的資料夾
        if os.path.isdir('library\\' + d):               # 獲得資料夾
            faces = glob.glob('library\\' + d + '\\*.jpg')  # 資料夾中所有人臉
            for face in faces:                          # 讀取人臉
                img = cv2.imread(face, cv2.IMREAD_GRAYSCALE)
                faces_db.append(img)                    # 人臉存入串列
                labels.append(index)                    # 建立數值標籤
            nameList.append(d)                          # 將學號加入串列
            index += 1
    print(f"標籤名稱 = {nameList}")
    print(f"標籤序號 ={labels}")
    # 儲存人名串列，可在未來辨識人臉時使用
    f = open('library\\employee.txt', 'w')
    test = open('library\\' + name + "\\" + "test.txt", 'w')
    checkpng = 'library\\' + name
    checkpngfiles = os.listdir(checkpng)
    num_png = len(checkpngfiles)
    num_png -= 1
    num_png = str(num_png)
    test.write(num_png)
    f.write(','.join(nameList))
    f.close()
    test.close()
    print('建立人臉辨識資料庫')
    model = cv2.face.LBPHFaceRecognizer_create()        # 建立LBPH人臉辨識物件
    model.train(faces_db, np.array(labels))             # 訓練LBPH人臉辨識
    model.save('library\\deepmind.yml')                  # 儲存LBPH訓練數據
    messagebox.showinfo(f'提示',f'人臉辨識資料庫完成\n目前存在的學生:\n{nameList}')
    print(f'人臉辨識資料庫完成\n目前存在的學生:\n標籤名稱 = {nameList}')
    Name=name
    new()
#-------------------------------------------------------------------------------
def Identify():
    global NAME
    mp_face_detection = mp.solutions.face_detection   # 建立偵測方法
    mp_drawing = mp.solutions.drawing_utils           # 建立繪圖方法
    
    if not os.path.exists("library\\deepmind.yml"):
        messagebox.showerror( "Erro!","沒有人臉數據\n請點 [加入人臉模型]")
    model = cv2.face.LBPHFaceRecognizer_create()
    model.read('library\\deepmind.yml')                  # 讀取已訓練模型
    f = open('library\\employee.txt', 'r')               # 開啟姓名標籤
    names = f.readline().split(',')                     # 將姓名存於串列

    def gocap():
        cap = cv2.VideoCapture(0)
        while(cap.isOpened()):                              # 如果開啟攝影機成功
            with mp_face_detection.FaceDetection(             # 開始偵測人臉
            model_selection=1, min_detection_confidence=0.5) as face_detection:
                while True:
                    ret, img = cap.read()
                    if not ret:
                        print("Cannot receive frame")
                        break
                    img2 = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)   # 將 BGR 顏色轉換成 RGB
                    results = face_detection.process(img2)        # 偵測人臉

                    if results.detections:
                        for detection in results.detections:
                            bboxC = detection.location_data.relative_bounding_box
                            ih, iw, ic = img.shape
                            bbox = (int(bboxC.xmin * iw), int(bboxC.ymin * ih), int(bboxC.width * iw), int(bboxC.height * ih))
                            cv2.rectangle(img, bbox, (255,255,255), 1)
                    cv2.imshow('oxxostudio', img)
                    x = int(bboxC.xmin * iw)
                    y = int(bboxC.ymin * ih)
                    w = int(bboxC.width * iw)
                    h = int(bboxC.height * ih)
                    key = cv2.waitKey(200)
                    if ret == True:       
                        print("已偵測到臉, 可按住空格鍵進行拍照檢測")
                        if key == ord(" "):          # 按 空格
                            imageCrop = img[y:y+h,x:x+w]                    # 裁切
                            imageResize = cv2.resize(imageCrop,(160,160))   # 重製大小
                            cv2.imwrite("library\\face.jpg", imageResize)    # 將測試人臉存檔
                            break
            cap.release()                                       # 關閉攝影機
            cv2.destroyAllWindows()
            gray = cv2.imread("library\\face.jpg", cv2.IMREAD_GRAYSCALE)
            val = model.predict(gray)
    gocap()
    gray = cv2.imread("library\\face.jpg", cv2.IMREAD_GRAYSCALE)
    val = model.predict(gray)
    if val[1] < 50:                                     #人臉辨識成功
        messagebox.showinfo( "歡迎DYU資工學生", f"歡迎DYU資工學生: {names[val[0]]}\n匹配值是: {val[1]:6.2f}")
        NAME=names[val[0]]
        record()
    #elif val[1] >50 and val[1] < 80:
        #test22 = tk.messagebox.askquestion('提示!',f"你跟: {names[val[0]]}匹配值是: {val[1]:6.2f} | 非常接近")
        #if test22 == 'yes':
            #NAME=names[val[0]]
            #path = f'library\\' + NAME +'\\' + 'test.txt'
            #readnub = open(path, 'r')
            #add = readnub.read()
            #cover = open(path, 'r+')
            #add = int(add)
            #add = add + 1
            #add = str(add)
            #cover.write(add)
            #print(add)
            #f1 = 'library\\face.jpg'    # 欲複製的檔案
            #f2 = 'library\\' + NAME + '\\' + NAME + add + '.jpg'  # 存檔的位置與檔案名稱
            #shutil.copyfile(f1,f2)   # 複製檔案
            #messagebox.showinfo( "歡迎DYU資工學生", f"歡迎DYU資工學生: {names[val[0]]}")
            #record()
        #else:
            #nomod = tk.messagebox.askquestion('提示!',"點名失敗!\n如果是新加選同學請點[加入人臉模型] 或者按下方的 [是] 錄製\n如果不是請重新進行識別")
            #if nomod == 'yes':
                #Train()
    else:
        nomod = tk.messagebox.askquestion('提示!',"點名失敗!\n如果是新加選同學請點[加入人臉模型] 或者按下方的 [是] 錄製\n如果不是請重新進行識別")
        if nomod == 'yes':
            Train()

def oas():
 global root
 root = tk.Tk()
 style = ThemedStyle(root)
 style.set_theme("breeze")
 root.geometry("1600x500")
 root.title('點名紀錄')
 qq="test.xlsx"
 df = pd.read_excel(qq,header=0)
 cols = list(df.columns)
 tree = ttk.Treeview(root)
 tree.pack()
 tree["columns"] = cols
 for i in cols:
    tree.column(i,anchor="center",width=10)
    tree.heading(i,text=i,anchor='center')
 for index, row in df.iterrows():
    tree.insert("",'end',text = index,values=list(row))
 tree.place(relx=0,rely=0.1,relheight=0.7,relwidth=1)
def new():
    tonow = datetime.datetime.now()
    yyyy=tonow.year
    mm=tonow.month
    dd=tonow.day
    week= (int(datetime.date(yyyy, mm, dd).strftime("%W")) -
           int(datetime.date(yyyy, mm, 1).strftime("%W")))*2
    today = date.today()
    d1 = today.strftime("%d")
    days=calendar.monthrange(yyyy,mm)[1]
    excel()
    pp="test.xlsx"
    wb = load_workbook(pp)
    sheet = wb.worksheets[0]# 取得第一個工作表
    wb.active = 0
    ws = wb.active
    ####################################修改
    a=str(ws.max_row+1) #新增下一位資料
    ws['A'+str(a)].value = name
    dayA=chr(65+int(d1)-week)
    sheet[dayA+str(a)].value =  '●'
    wb.save(pp) 
def record():
    tonow = datetime.datetime.now()
    yyyy=tonow.year
    mm=tonow.month
    dd=tonow.day
    week= (int(datetime.date(yyyy, mm, dd).strftime("%W")) -
           int(datetime.date(yyyy, mm, 1).strftime("%W")))*2
    pp="test.xlsx"
    excel()
    ####################################修改
    wb = load_workbook(pp)
    sheet = wb.worksheets[0]# 取得第一個工作表
    ####################################出席紀錄
    wb.active = 0
    ws = wb.active
    a=str(ws.max_row+1) #新增下一筆點名記錄
    i=1
    while i<int(a):
        if sheet['A'+str(i)].value==NAME:
            dayA=chr(65+int(d1)-week)
            sheet[dayA+str(i)].value =  '●'
            wb.save(pp) 
            break
        else:
            i+=1
class Toplevel1:
    def __init__(self, top=None):
        '''This class configures and populates the toplevel window.
           top is the toplevel containing window.'''

        top.geometry("566x265+723+328")
        top.minsize(566,265)
        top.maxsize(566,265)
        top.resizable(1,  1)
        top.title("人臉點名系統")
        top.configure(background="#ffffff", highlightbackground="#d9d9d9", highlightcolor="black")

        self.top = top

        self.Button1 = tk.Button(self.top)
        self.Button1.place(relx=0.004, rely=0.011, height=260, width=280)
        self.Button1.configure(activebackground="#e7e7e7", activeforeground="black", background="#efefef", compound='left', 
                                cursor="hand2", disabledforeground="#a3a3a3", foreground="#000000", highlightbackground="#d9d9d9", 
                                highlightcolor="black", pady="0", relief="ridge", text='''點名系統''', font=('微軟正黑體 Light', '14'), 
                                command=Identify)

        self.Button1_1 = tk.Button(self.top)
        self.Button1_1.place(relx=0.5, rely=0.011, height=130, width=280)
        self.Button1_1.configure(activebackground="#e7e7e7", activeforeground="black", background="#efefef", compound='left', 
                                    cursor="hand2", disabledforeground="#a3a3a3", foreground="#000000", highlightbackground="#d9d9d9", 
                                    highlightcolor="black", pady="0", relief="ridge", text='''加入人臉模型''', font=('微軟正黑體 Light', '12'), 
                                    command=Train)

        self.Button1_1_1 = tk.Button(self.top)
        self.Button1_1_1.place(relx=0.5, rely=0.506, height=130, width=280)
        self.Button1_1_1.configure(activebackground="#e7e7e7", activeforeground="black", background="#efefef", compound='left', 
                                    cursor="hand2", disabledforeground="#a3a3a3", foreground="#000000", highlightbackground="#d9d9d9", 
                                    highlightcolor="black", pady="0", relief="ridge", text='''查看點名詳情''', font=('微軟正黑體 Light', '12'),
                                    command=oas)

if __name__ == '__main__':
    '''Main entry point for the application.'''
    global root
    root = tk.Tk()
    root.protocol( 'WM_DELETE_WINDOW' , root.destroy)
    # Creates a toplevel widget.
    global _top1, _w1
    _top1 = root
    _w1 = Toplevel1(_top1)
    root.mainloop() 
